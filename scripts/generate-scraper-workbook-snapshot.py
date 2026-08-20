#!/usr/bin/env python3
"""Build the server-only ATHRECS scraper workbook snapshot module.

Usage:
  python scripts/generate-scraper-workbook-snapshot.py /path/to/workbook.xlsx

The generated TypeScript contains a gzip-compressed JSON payload so the full
audit snapshot never enters the browser bundle.
"""

from __future__ import annotations

import base64
import csv
import gzip
import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, time, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[1]
REGISTRY_PATH = ROOT / "docs/source-registry/fixture-result-sources.csv"
OVERRIDES_PATH = ROOT / "docs/source-registry/workbook-event-overrides.json"
OUTPUT_PATH = ROOT / "src/lib/athrecs/scraper-workbook-snapshot.server.generated.ts"
PARTS_PATH = ROOT / "src/lib/athrecs/scraper-workbook-snapshot-parts.server.generated"
PART_SIZE = 60_000


def sheet_records(workbook, sheet_name: str):
    rows = workbook[sheet_name].iter_rows(values_only=True)
    for _ in range(3):
        next(rows)
    headers = next(rows)
    for values in rows:
        if not any(value is not None for value in values):
            continue
        yield dict(zip(headers, values))


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value).strip()


def number(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        converted = from_excel(value)
        return converted.date().isoformat() if isinstance(converted, datetime) else converted.isoformat()
    raw = text(value)
    return raw[:10] if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", raw) else ""


def iso_time(value) -> str:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0).isoformat(timespec="minutes")
    if isinstance(value, time):
        return value.replace(microsecond=0).isoformat(timespec="minutes")
    raw = text(value)
    return raw[:5] if re.fullmatch(r"\d{2}:\d{2}.*", raw) else ""


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")[:80]


def mapped_status(edition_status: str, entry_status: str) -> str:
    if edition_status == "Completed":
        return "Finished"
    if edition_status == "Cancelled":
        return "Closed"
    if entry_status == "Open":
        return "Open"
    if entry_status == "Sold Out":
        return "Closed"
    if entry_status == "Closed":
        return "Closed"
    return "TBC"


def mapped_sport(value: str) -> str:
    lowered = value.lower()
    for sport in (
        "Athletics",
        "Parkrun",
        "Cycling",
        "Swimming",
        "Triathlon",
        "Duathlon",
        "Aquathlon",
        "Aquabike",
        "Rowing",
        "OCR",
    ):
        if sport.lower() in lowered:
            return sport
    return "Running"


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Pass the scraper workbook path")
    workbook_path = Path(sys.argv[1]).resolve()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    with REGISTRY_PATH.open(newline="", encoding="utf-8-sig") as source:
        registry_rows = list(csv.DictReader(source))
    registry = {row["source_id"].strip(): row for row in registry_rows}
    enabled_sources = {
        source_id
        for source_id, row in registry.items()
        if text(row.get("enabled")).lower() in {"1", "true", "yes"}
    }
    override_document = json.loads(OVERRIDES_PATH.read_text(encoding="utf-8"))
    overrides = override_document.get("overrides", {})
    applied_overrides = set()

    races = {text(row["Race ID"]): row for row in sheet_records(workbook, "Races")}
    courses = defaultdict(list)
    for row in sheet_records(workbook, "Courses"):
        courses[text(row["Edition ID"])].append(row)

    issues = defaultdict(list)
    for row in sheet_records(workbook, "Review Queue"):
        if text(row["Review Status"]).lower() not in {"open", "pending", "needs review"}:
            continue
        issues[(text(row["Entity Type"]), text(row["Entity ID"]))].append(
            {
                "field": text(row["Field Name"]),
                "issueType": text(row["Issue Type"]),
                "priority": text(row["Priority"]),
                "sourceUrl": text(row["Source URL"]),
            }
        )

    register = list(sheet_records(workbook, "Workbook Register"))[-1]
    snapshot_id = text(register["Workbook ID"])
    created_value = register["Created At"]
    if isinstance(created_value, (int, float)):
        created_value = from_excel(created_value)
    if not isinstance(created_value, datetime):
        raise ValueError("Latest workbook register row has no valid Created At value")
    created_at = created_value.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

    candidates = []
    for edition in sheet_records(workbook, "Editions"):
        edition_id = text(edition["Edition ID"])
        race_id = text(edition["Race ID"])
        race = races.get(race_id, {})
        source_id = text(edition["Source ID"])
        event_name = text(edition["Race Name"] or race.get("Race Name"))
        event_date = iso_date(edition["Start Date"])
        source_url = text(edition["Source URL"])
        source_hash = text(edition["Source Row Hash"])
        override = overrides.get(edition_id)
        if override:
            if text(override.get("sourceId")) != source_id:
                raise ValueError(f"Override source mismatch for {edition_id}")
            if text(override.get("eventName")) != event_name:
                raise ValueError(f"Override event name mismatch for {edition_id}")
            if text(override.get("eventDate")) != event_date:
                raise ValueError(f"Override event date mismatch for {edition_id}")
            applied_overrides.add(edition_id)
        override_fields = override.get("fields", {}) if override else {}
        resolved_issue_fields = set(override.get("resolvedIssueFields", [])) if override else set()
        edition_issues = issues[("Edition", edition_id)]
        race_issues = issues[("Race", race_id)]
        all_issues = [
            issue
            for issue in race_issues + edition_issues
            if issue["field"] not in resolved_issue_fields
        ]
        high_issues = [issue for issue in all_issues if issue["priority"] == "High"]

        distances = []
        for course in courses[edition_id]:
            label = text(course["Distance Text"] or course["Distance Type"] or course["Course Name"])
            if not label:
                continue
            distances.append(
                {
                    "code": label,
                    "km": number(course["Distance KM"]) or 0,
                    "surface": text(course["Surface"]),
                    "sourceUrl": text(course["Source URL"]),
                }
            )
        if not distances and text(edition["Distance Text"]):
            distances.append(
                {
                    "code": text(edition["Distance Text"]),
                    "km": 0,
                    "surface": text(edition["Surface"]),
                    "sourceUrl": source_url,
                }
            )
        deduped_distances = []
        seen_distance_codes = set()
        for distance in distances:
            key = distance["code"].lower()
            if key not in seen_distance_codes:
                seen_distance_codes.add(key)
                deduped_distances.append(distance)

        if "distances" in override_fields:
            deduped_distances = override_fields["distances"]

        country = text(override_fields.get("country") or edition["Country"] or race.get("Country"))
        official_website = text(
            override_fields.get("website")
            or edition["Official Website URL"]
            or race.get("Official Website URL")
        )
        edition_status = text(edition["Status"])
        entry_status = text(edition["Entry Status"])
        block_reasons = []
        if source_id not in registry:
            block_reasons.append("source_not_in_current_registry")
        elif source_id not in enabled_sources:
            block_reasons.append("source_disabled")
        if edition_status not in {"Confirmed", "Completed"}:
            block_reasons.append(f"edition_status_{slugify(edition_status) or 'missing'}")
        for field, present in (
            ("event_name", bool(event_name)),
            ("event_date", bool(event_date)),
            ("country", bool(country)),
            ("distance", bool(deduped_distances)),
            ("source_url", bool(source_url)),
            ("source_row_hash", bool(source_hash)),
        ):
            if not present:
                block_reasons.append(f"missing_{field}")
        if high_issues:
            block_reasons.append("open_high_priority_review_item")

        event_slug = text(race.get("Slug")) or slugify(event_name)
        fingerprint_input = "|".join(
            [
                re.sub(r"[^a-z0-9]+", "", event_name.lower()),
                event_date,
                country.lower(),
                "|".join(sorted(distance["code"].lower() for distance in deduped_distances)),
            ]
        )
        fingerprint = hashlib.sha256(fingerprint_input.encode()).hexdigest()
        candidate = {
            "id": f"{snapshot_id}:{edition_id}",
            "sourceRaceId": race_id,
            "sourceEditionId": edition_id,
            "sourceId": source_id,
            "sourceUrl": source_url,
            "sourceRowHash": source_hash,
            "fingerprint": fingerprint,
            "eventName": event_name,
            "eventSlug": event_slug,
            "eventDate": event_date,
            "reviewItemCount": len(all_issues),
            "highIssueCount": len(high_issues),
            "blockReasons": sorted(set(block_reasons)),
            "payload": {
                "raceId": race_id,
                "editionId": edition_id,
                "name": event_name,
                "slug": event_slug,
                "description": text(race.get("Description")),
                "sport": mapped_sport(text(race.get("Sport") or edition["Race Type"])),
                "country": country,
                "county": text(edition["County Region"] or race.get("County Region")),
                "stateProvince": text(
                    edition["State Province"] or race.get("State Province")
                ),
                "city": text(override_fields.get("city") or edition["City"] or race.get("City")),
                "venue": text(
                    override_fields.get("venue")
                    or edition["Venue Name"]
                    or race.get("Venue Name")
                ),
                "surface": text(edition["Surface"] or race.get("Primary Surface")) or "Road",
                "organiser": text(race.get("Organiser Name")),
                "website": official_website,
                "date": event_date,
                "startTime": iso_time(edition["Start Time Local"]),
                "status": text(override_fields.get("status"))
                or mapped_status(edition_status, entry_status),
                "editionStatus": edition_status,
                "entryStatus": text(override_fields.get("entryStatus")) or entry_status,
                "entryUrl": text(edition["Entry URL"]),
                "sourceId": source_id,
                "sourceUrl": source_url,
                "sourceRowHash": source_hash,
                "confidence": text(edition["Confidence"]),
                "qualityScore": number(race.get("Quality Score")) or 0,
                "distances": deduped_distances,
                "issues": all_issues,
                "enrichment": (
                    {
                        "checkedAt": text(override_document.get("checkedAt")),
                        "evidenceUrl": text(override.get("evidenceUrl")),
                        "resolvedIssueFields": sorted(resolved_issue_fields),
                        "note": text(override.get("note")),
                    }
                    if override
                    else None
                ),
            },
        }
        candidates.append(candidate)

    unused_overrides = sorted(set(overrides) - applied_overrides)
    if unused_overrides:
        raise ValueError(f"Workbook overrides did not match an edition: {unused_overrides}")

    payload = {
        "snapshotId": snapshot_id,
        "snapshotCreatedAt": created_at,
        "workbookRunId": text(register["Run ID"]),
        "sourceCount": int(register["Source Count"]),
        "raceCount": int(register["Race Count"]),
        "editionCount": int(register["Edition Count"]),
        "courseCount": int(register["Course Count"]),
        "reviewItemCount": int(register["Review Item Count"]),
        "workbookStatus": text(register["Status"]),
        "candidates": candidates,
    }
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    encoded = base64.b64encode(gzip.compress(raw, compresslevel=9, mtime=0)).decode("ascii")
    PARTS_PATH.mkdir(parents=True, exist_ok=True)
    for old_part in PARTS_PATH.glob("part-*.ts"):
        old_part.unlink()
    parts = [encoded[index : index + PART_SIZE] for index in range(0, len(encoded), PART_SIZE)]
    imports = []
    names = []
    for index, part in enumerate(parts):
        name = f"part{index}"
        filename = f"part-{index:03d}.ts"
        (PARTS_PATH / filename).write_text(
            "// Generated scraper snapshot fragment. Do not edit.\n"
            f'export default "{part}";\n',
            encoding="utf-8",
        )
        imports.append(
            f'import {name} from "./scraper-workbook-snapshot-parts.server.generated/{filename[:-3]}";'
        )
        names.append(name)
    output = "\n".join(
        [
            "// Generated by scripts/generate-scraper-workbook-snapshot.py. Do not edit.",
            *imports,
            "",
            "export const SCRAPER_WORKBOOK_SNAPSHOT_GZIP_BASE64 =",
            f"  [{', '.join(names)}].join(\"\");",
            "",
        ]
    )
    OUTPUT_PATH.write_text(output, encoding="utf-8")
    eligible = sum(not candidate["blockReasons"] for candidate in candidates)
    print(
        json.dumps(
            {
                "output": str(OUTPUT_PATH),
                "candidates": len(candidates),
                "eligibleBeforeCatalogueDedupe": eligible,
                "compressedBytes": len(encoded),
                "generatedParts": len(parts),
                "snapshotId": snapshot_id,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
